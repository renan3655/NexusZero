# -*- coding: utf-8 -*-
"""
ANÁLISE DE JOGOS 0x0 EM TEMPO REAL - v2.2
Desenvolvido por: Renan Quintanilha Marques 
Última atualização: 06/04/2025  
"""

import json
import re
from httpx import get, RequestError
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
from telegram_notifier import enviar_notificacao  # Importação da nova função
import time  # Para controle de intervalo de tempo

# ====================== CONFIGURAÇÕES ======================
HEADER_COLOR = "2A629A"  # Azul mais moderno
FILE_NAME = "Relatorio_Jogos_0x0.xlsx"
MIN_MINUTES = 20  # Tempo mínimo para análise
MAX_MINUTES = 120  # Tempo máximo considerado
SHOW_LIMIT = 10  # Quantidade de jogos exibidos no console

# Variável global para jogos já notificados
jogos_notificados = set()

# ====================== FUNÇÕES PRINCIPAIS ======================
class AnalisadorJogos:
    @staticmethod
    def obter_dados():
        """Obtém dados com tratamento de erros reforçado"""
        try:
            print("🔍 Conectando ao OneFootball...", end=" ", flush=True)
            response = get(
                'https://onefootball.com/pt-br/jogos?only_live=true',
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
                    'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7'
                },
                timeout=20
            )
            response.raise_for_status()

            if match := re.search(r'\{"props".+\}', response.text, re.DOTALL):
                print("✅ Dados obtidos com sucesso!")
                return json.loads(match.group())
            
            print("⚠️ Estrutura do site alterada!")
            return None

        except RequestError as e:
            print(f"🔴 Falha na conexão: {str(e)}")
            return None
        except Exception as e:
            print(f"⚠️ Erro inesperado: {type(e).__name__} - {str(e)}")
            return None

    @staticmethod
    def processar_tempo(tempo_str):
        """Processa o tempo com inteligência para acréscimos"""
        if not tempo_str or tempo_str.lower() in ['não iniciado', 'intervalo']:
            return 0, tempo_str
        
        try:
            digitos = ''.join(c for c in tempo_str if c.isdigit())
            
            if len(digitos) > 2:  # Caso 90+2 (902)
                tempo_regulamentar = int(digitos[:2])
                acrescimos = int(digitos[2:])
                total = min(90 + acrescimos, MAX_MINUTES)
                return total, f"{tempo_regulamentar}+{acrescimos}'"
            
            minutos = int(digitos) if digitos else 0
            return min(minutos, MAX_MINUTES), f"{minutos}'"
            
        except ValueError:
            return 0, tempo_str

    @staticmethod
    def gerar_relatorio(jogos):
        """Cria relatório Excel profissional"""
        if not jogos:
            print("📭 Nenhum dado para gerar relatório")
            return False

        try:
            print("\n📊 Gerando relatório Excel...")
            wb = Workbook()
            planilha = wb.active
            planilha.title = "Jogos 0x0"
            
            # Estilos premium
            estilo_cabecalho = Font(bold=True, color="FFFFFF", size=11)
            preenchimento = PatternFill(
                start_color=HEADER_COLOR,
                end_color=HEADER_COLOR,
                fill_type="solid"
            )
            bordas = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Cabeçalhos
            colunas = [
                'Competição', 'Time Casa', 'Placar Casa',
                'Time Visitante', 'Placar Visitante',
                'Tempo Jogo', 'Minutos', 'Status',
                'Data', 'Hora'
            ]
            planilha.append(colunas)
            
            # Dados
            for jogo in jogos:
                planilha.append([jogo[col] for col in colunas])
            
            # Formatação automática
            for idx, coluna in enumerate(colunas, 1):
                letra = get_column_letter(idx)
                
                # Cabeçalho
                celula = planilha[f"{letra}1"]
                celula.font = estilo_cabecalho
                celula.fill = preenchimento
                celula.alignment = Alignment(horizontal='center')
                celula.border = bordas
                
                # Ajuste de largura
                planilha.column_dimensions[letra].width = max(len(coluna) + 2, 15)
                
                # Dados
                for linha in range(2, len(jogos) + 2):
                    celula = planilha[f"{letra}{linha}"]
                    celula.alignment = Alignment(horizontal='center')
                    celula.border = bordas
            
            planilha.freeze_panes = 'A2'
            planilha.auto_filter.ref = planilha.dimensions
            wb.save(FILE_NAME)
            
            print(f"💾 Relatório salvo como '{FILE_NAME}'")
            return True
            
        except Exception as e:
            print(f"⛔ Falha ao gerar relatório: {type(e).__name__}")
            return False

# ====================== EXECUÇÃO ======================
if __name__ == "__main__":
    try:
        print(f"\n⚽ NEXUS ZERO - Desenvolvido por Renan Quintanilha Marques\n")
        
        # Loop contínuo para execução automática a cada 15 minutos
        while True:
            print("\n🔄 Iniciando nova análise...\n")
            
            # Coleta de dados
            dados = AnalisadorJogos.obter_dados()
            if not dados:
                print("⚠️ Dados não obtidos, tentando novamente na próxima execução...")
                dados = {"props": {"pageProps": {"containers": []}}}  # Dados vazios para evitar falhas
            
            # Processamento
            jogos_validos = []
            for jogo in dados.get('props', {}).get('pageProps', {}).get('containers', []):
                try:
                    if cards := jogo.get('type', {}).get('fullWidth', {}).get('component', {}).get('contentType', {}).get('matchCardsList', {}):
                        for partida in cards.get('matchCards', []):
                            try:
                                # Extração segura
                                time_casa = partida['homeTeam']['name']
                                placar_casa = int(partida['homeTeam']['score'].split()[0]) if partida['homeTeam']['score'] else 0
                                time_fora = partida['awayTeam']['name']
                                placar_fora = int(partida['awayTeam']['score'].split()[0]) if partida['awayTeam']['score'] else 0
                                
                                minutos, tempo_formatado = AnalisadorJogos.processar_tempo(partida.get('timePeriod'))
                                
                                if placar_casa == 0 and placar_fora == 0 and 20 <= minutos <= 40:
                                    jogo = {
                                        'Competição': partida['trackingEvents'][0]['typedServerParameter']['competition']['value'],
                                        'Time Casa': time_casa,
                                        'Placar Casa': placar_casa,
                                        'Time Visitante': time_fora,
                                        'Placar Visitante': placar_fora,
                                        'Tempo Jogo': tempo_formatado,
                                        'Minutos': minutos,
                                        'Status': 'Em Andamento' if minutos > 0 else 'Pré-Jogo',
                                        'Data': datetime.now().strftime('%d/%m/%Y'),
                                        'Hora': datetime.now().strftime('%H:%M:%S')
                                    }
                                    jogos_validos.append(jogo)
                                    
                                    # ENVIA NOTIFICAÇÃO PARA O TELEGRAM
                                    if jogo['Time Casa'] + jogo['Time Visitante'] not in jogos_notificados:
                                        try:
                                            enviar_notificacao(
                                                time_casa=jogo['Time Casa'],
                                                time_fora=jogo['Time Visitante'],
                                                competicao=jogo['Competição'],
                                                tempo=jogo['Tempo Jogo']
                                            )
                                            # Marque o jogo como notificado
                                            jogos_notificados.add(jogo['Time Casa'] + jogo['Time Visitante'])
                                        except Exception as e:
                                            print(f"⚠️ Falha no Telegram: {e}")
                            except Exception:
                                continue
                except Exception:
                    continue
            
            # Resultados
            print(f"\n📋 Total de jogos válidos: {len(jogos_validos)}")
            for i, jogo in enumerate(jogos_validos[:SHOW_LIMIT], 1):
                print(f"{i}. {jogo['Time Casa']} {jogo['Placar Casa']}×{jogo['Placar Visitante']} {jogo['Time Visitante']} | {jogo['Tempo Jogo']}")
                
            
                        # Relatório
            AnalisadorJogos.gerar_relatorio(jogos_validos)
            
            # Calcula o horário da próxima execução
            hora_atual = datetime.now()
            proxima_execucao = hora_atual + timedelta(minutes=15)
            print(f"\n⏳ Aguardando próxima execução às {proxima_execucao.strftime('%H:%M:%S')}... Pressione Ctrl+C para sair\n")
            
            # Aguarda 15 minutos antes de continuar o loop
            time.sleep(900)  # 15 minutos em segundos
        
    except KeyboardInterrupt:
        print("\n⏹️ Monitoramento interrompido pelo usuário")
    except Exception as e:
        print(f"\n❌ ERRO: {type(e).__name__} - {str(e)}")

